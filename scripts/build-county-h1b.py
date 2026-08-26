#!/usr/bin/env python3
"""Convert DOL FY2025 LCA workbooks into a compact county H-1B JSON file."""

from __future__ import annotations

import csv
import json
import math
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


CERTIFIED_STATUSES = {"CERTIFIED"}
UNIT_MULTIPLIERS = {
    "HOUR": 2080,
    "WEEK": 52,
    "BI-WEEKLY": 26,
    "BIWEEKLY": 26,
    "MONTH": 12,
    "YEAR": 1,
}


def main() -> None:
    if len(sys.argv) != 8:
        raise SystemExit(
            "Usage: build-county-h1b.py <LCA xlsx> <worksites xlsx> "
            "<county GeoJSON> <national ZIP relationship> <CT ZIP relationship> "
            "<output JSON> <validation JSON>"
        )
    (
        lca_paths_arg,
        worksites_path,
        county_path,
        national_zip_path,
        ct_zip_path,
        output_path,
        validation_path,
    ) = map(Path, sys.argv[1:])

    geography = Geography(county_path, national_zip_path, ct_zip_path)
    lca_paths = [Path(value) for value in str(lca_paths_arg).split(",")]
    cases, case_stats = read_h1b_cases(lca_paths)
    counties, worksite_stats, unmatched = aggregate_worksites(
        worksites_path, cases, geography
    )
    output = build_output(counties, geography.county_names)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, separators=(",", ":")) + "\n")

    validation = {
        "sourceCases": case_stats,
        "worksites": worksite_stats,
        "outputCountyCount": len(output["counties"]),
        "topUnmatchedLocations": [
            {"location": key, "workerPlacements": value}
            for key, value in unmatched.most_common(30)
        ],
    }
    validation_path.write_text(json.dumps(validation, indent=2) + "\n")
    print(
        f"Wrote {len(output['counties']):,} counties; "
        f"mapped {worksite_stats['mappedCertifiedWorkerPlacements']:,} of "
        f"{worksite_stats['certifiedWorkerPlacements']:,} certified placements."
    )


def read_h1b_cases(paths: list[Path]):
    cases = {}
    statuses = Counter()
    for path in paths:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        column = {name: index for index, name in enumerate(header)}
        required = [
            "CASE_NUMBER",
            "CASE_STATUS",
            "VISA_CLASS",
            "JOB_TITLE",
            "SOC_CODE",
            "SOC_TITLE",
            "FULL_TIME_POSITION",
            "TOTAL_WORKER_POSITIONS",
            "EMPLOYER_NAME",
        ]
        require_columns(column, required, path)
        for row in rows:
            if clean(row[column["VISA_CLASS"]]).upper() != "H-1B":
                continue
            case_number = clean(row[column["CASE_NUMBER"]])
            if not case_number:
                continue
            status = clean(row[column["CASE_STATUS"]]).upper()
            statuses[status or "MISSING"] += 1
            cases[case_number] = {
                "status": status,
                "certified": status in CERTIFIED_STATUSES,
                "employer": clean(row[column["EMPLOYER_NAME"]])
                or "Unknown employer",
                "socCode": clean(row[column["SOC_CODE"]]),
                "occupation": clean(row[column["SOC_TITLE"]])
                or clean(row[column["JOB_TITLE"]])
                or "Unknown occupation",
                "fullTime": clean(row[column["FULL_TIME_POSITION"]]).upper() == "Y",
                "totalWorkerPositions": nonnegative_int(
                    row[column["TOTAL_WORKER_POSITIONS"]]
                ),
            }
        workbook.close()
    final_statuses = Counter(case["status"] or "MISSING" for case in cases.values())
    return cases, {
        "quarterFiles": [path.name for path in paths],
        "h1bApplications": len(cases),
        "certifiedApplications": sum(case["certified"] for case in cases.values()),
        "certifiedWorkerPositions": sum(
            case["totalWorkerPositions"] for case in cases.values() if case["certified"]
        ),
        "allWorkerPositions": sum(
            case["totalWorkerPositions"] for case in cases.values()
        ),
        "finalStatusCounts": dict(sorted(final_statuses.items())),
        "quarterRowStatusCounts": dict(sorted(statuses.items())),
    }


def aggregate_worksites(path: Path, cases: dict, geography: "Geography"):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    column = {name: index for index, name in enumerate(header)}
    required = [
        "CASE_NUMBER",
        "WORKSITE_WORKERS",
        "SECONDARY_ENTITY",
        "WORKSITE_CITY",
        "WORKSITE_COUNTY",
        "WORKSITE_STATE",
        "WORKSITE_POSTAL_CODE",
        "WAGE_RATE_OF_PAY_FROM",
        "WAGE_UNIT_OF_PAY",
        "PREVAILING_WAGE",
        "PW_UNIT_OF_PAY",
    ]
    require_columns(column, required, path)

    counties = defaultdict(new_county)
    unmatched = Counter()
    stats = Counter()
    for row in rows:
        case_number = clean(row[column["CASE_NUMBER"]])
        case = cases.get(case_number)
        if not case:
            continue
        stats["h1bWorksiteRows"] += 1
        workers = nonnegative_int(row[column["WORKSITE_WORKERS"]])
        stats["h1bWorkerPlacements"] += workers
        if case["certified"]:
            stats["certifiedWorksiteRows"] += 1
            stats["certifiedWorkerPlacements"] += workers

        state = clean(row[column["WORKSITE_STATE"]])
        county_name = clean(row[column["WORKSITE_COUNTY"]])
        city = clean(row[column["WORKSITE_CITY"]])
        postal_code = zip5(row[column["WORKSITE_POSTAL_CODE"]])
        geoid = geography.find(state, county_name, postal_code)
        if not geoid:
            stats["unmatchedWorksiteRows"] += 1
            stats["unmatchedWorkerPlacements"] += workers
            if case["certified"]:
                stats["unmatchedCertifiedWorkerPlacements"] += workers
            unmatched[f"{state}|{county_name}|{city}|{postal_code}"] += workers
            continue
        stats["mappedWorksiteRows"] += 1
        stats["mappedWorkerPlacements"] += workers
        if case["certified"]:
            stats["mappedCertifiedWorkerPlacements"] += workers

        county = counties[geoid]
        county["applications"].add(case_number)
        county["allWorkerPlacements"] += workers
        if not case["certified"]:
            continue
        county["certifiedApplications"].add(case_number)
        county["certifiedWorkerPlacements"] += workers
        if case["fullTime"]:
            county["fullTimeWorkerPlacements"] += workers
        if clean(row[column["SECONDARY_ENTITY"]]).upper() == "Y":
            county["secondaryEntityWorkerPlacements"] += workers
        county["employers"][case["employer"]] += workers
        occupation_key = f"{case['socCode']}|{case['occupation']}"
        county["occupations"][occupation_key] += workers

        offered = annualize(
            row[column["WAGE_RATE_OF_PAY_FROM"]],
            row[column["WAGE_UNIT_OF_PAY"]],
        )
        prevailing = annualize(
            row[column["PREVAILING_WAGE"]], row[column["PW_UNIT_OF_PAY"]]
        )
        if offered is not None and workers:
            county["offeredWageWeightedSum"] += offered * workers
            county["offeredWageWeight"] += workers
        if prevailing is not None and workers:
            county["prevailingWageWeightedSum"] += prevailing * workers
            county["prevailingWageWeight"] += workers
        if offered is not None and prevailing not in (None, 0) and workers:
            county["wagePremiumWeightedSum"] += ((offered / prevailing) - 1) * workers
            county["wagePremiumWeight"] += workers

    workbook.close()
    return counties, dict(sorted(stats.items())), unmatched


def new_county():
    return {
        "applications": set(),
        "certifiedApplications": set(),
        "allWorkerPlacements": 0,
        "certifiedWorkerPlacements": 0,
        "fullTimeWorkerPlacements": 0,
        "secondaryEntityWorkerPlacements": 0,
        "offeredWageWeightedSum": 0.0,
        "offeredWageWeight": 0,
        "prevailingWageWeightedSum": 0.0,
        "prevailingWageWeight": 0,
        "wagePremiumWeightedSum": 0.0,
        "wagePremiumWeight": 0,
        "employers": Counter(),
        "occupations": Counter(),
    }


def build_output(counties: dict, county_names: dict):
    records = {}
    for geoid in sorted(counties):
        county = counties[geoid]
        top_employers = [
            {"name": name, "workerPlacements": workers}
            for name, workers in county["employers"].most_common(5)
        ]
        top_occupations = []
        for key, workers in county["occupations"].most_common(5):
            soc_code, title = key.split("|", 1)
            top_occupations.append(
                {"socCode": soc_code, "title": title, "workerPlacements": workers}
            )
        records[geoid] = {
            "name": county_names[geoid],
            "applications": len(county["applications"]),
            "certifiedApplications": len(county["certifiedApplications"]),
            "allWorkerPlacements": county["allWorkerPlacements"],
            "certifiedWorkerPlacements": county["certifiedWorkerPlacements"],
            "fullTimeWorkerPlacements": county["fullTimeWorkerPlacements"],
            "secondaryEntityWorkerPlacements": county[
                "secondaryEntityWorkerPlacements"
            ],
            "averageOfferedAnnualWage": weighted_average(
                county["offeredWageWeightedSum"], county["offeredWageWeight"]
            ),
            "averagePrevailingAnnualWage": weighted_average(
                county["prevailingWageWeightedSum"],
                county["prevailingWageWeight"],
            ),
            "averageWagePremiumPercent": weighted_average(
                county["wagePremiumWeightedSum"] * 100,
                county["wagePremiumWeight"],
                digits=2,
            ),
            "topEmployers": top_employers,
            "topOccupations": top_occupations,
        }
    return {
        "id": "h1b-certified-worker-placements",
        "label": "Certified H-1B worker placements",
        "description": (
            "Worker placements listed at county worksites on certified FY2025 "
            "H-1B Labor Condition Applications"
        ),
        "vintage": "FY2025",
        "period": "2024-10-01 through 2025-09-30",
        "source": {
            "label": "U.S. Department of Labor FY2025 LCA disclosure data",
            "url": "https://www.dol.gov/agencies/eta/foreign-labor/performance",
        },
        "caveat": (
            "A certified LCA is a wage attestation, not a USCIS petition approval or proof "
            "that a job was filled. Worker counts are requested worksite placements and "
            "may not represent unique people. Wages are employer-reported minimum offered "
            "rates, annualized and weighted by certified worksite placements. County "
            "assignment uses the reported worksite county with Census ZIP relationships "
            "as a fallback."
        ),
        "counties": records,
    }


class Geography:
    def __init__(self, county_path: Path, national_zip_path: Path, ct_zip_path: Path):
        geojson = json.loads(county_path.read_text())
        self.valid_geoids = set()
        self.county_names = {}
        self.geoid_states = {}
        state_names = {}
        candidates = defaultdict(set)
        for feature in geojson["features"]:
            properties = feature["properties"]
            geoid = properties["GEOID"]
            state = properties["STUSPS"]
            self.valid_geoids.add(geoid)
            self.county_names[geoid] = (
                f"{properties['NAMELSAD']}, {properties['STATE_NAME']}"
            )
            self.geoid_states[geoid] = state
            state_names[normalize(properties["STATE_NAME"])] = state
            state_names[normalize(state)] = state
            for name in (properties["NAME"], properties["NAMELSAD"]):
                for alias in name_aliases(name):
                    candidates[f"{state}|{alias}"].add(geoid)
        self.state_names = state_names
        self.county_aliases = {
            key: next(iter(geoids))
            for key, geoids in candidates.items()
            if len(geoids) == 1
        }
        self.zip_counties = largest_area_zip_match(
            national_zip_path,
            "GEOID_ZCTA5_20",
            "GEOID_COUNTY_20",
            lambda geoid: geoid,
            self.valid_geoids,
        )
        ct_matches = largest_area_zip_match(
            ct_zip_path,
            "GEOID_ZCTA5_20",
            "GEOID_COUSUB_22",
            lambda geoid: geoid[:5],
            self.valid_geoids,
        )
        self.zip_counties.update(ct_matches)

    def find(self, state_value: str, county_value: str, postal_code: str):
        state = self.state_names.get(normalize(state_value))
        zip_geoid = self.zip_counties.get(postal_code)
        name_geoid = None
        if state and county_value:
            for alias in name_aliases(county_value):
                name_geoid = self.county_aliases.get(f"{state}|{alias}")
                if name_geoid:
                    break
        if state == "CT":
            return zip_geoid or name_geoid
        if name_geoid:
            return name_geoid
        if zip_geoid and (not state or state == self.geoid_states[zip_geoid]):
            return zip_geoid
        return None


def largest_area_zip_match(
    path: Path, zip_column: str, geoid_column: str, normalize_geoid, valid_geoids
):
    candidates = {}
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter="|")
        for row in reader:
            postal_code = row[zip_column]
            geoid = normalize_geoid(row[geoid_column])
            if not postal_code or geoid not in valid_geoids:
                continue
            area = float(row["AREALAND_PART"] or 0) + float(row["AREAWATER_PART"] or 0)
            current = candidates.get(postal_code)
            if current is None or area > current[1]:
                candidates[postal_code] = (geoid, area)
    return {postal_code: geoid for postal_code, (geoid, _) in candidates.items()}


def name_aliases(value: str):
    normalized = normalize(value)
    aliases = {normalized}
    for suffix in (
        "CITYANDBOROUGH",
        "CENSUSAREA",
        "MUNICIPALITY",
        "BOROUGH",
        "PARISH",
        "COUNTY",
    ):
        if normalized.endswith(suffix):
            aliases.add(normalized[: -len(suffix)])
    for alias in list(aliases):
        if alias.startswith("SAINT"):
            aliases.add("ST" + alias[5:])
        if alias.startswith("ST"):
            aliases.add("SAINT" + alias[2:])
    return aliases


def normalize(value: str):
    text = unicodedata.normalize("NFD", clean(value))
    return "".join(character for character in text if character.isalnum()).upper()


def clean(value):
    return "" if value is None else str(value).strip()


def zip5(value):
    if value is None:
        return ""
    text = clean(value).split(".", 1)[0]
    digits = "".join(character for character in text if character.isdigit())
    return digits[:5].zfill(5) if digits else ""


def nonnegative_int(value):
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def annualize(value, unit):
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    multiplier = UNIT_MULTIPLIERS.get(clean(unit).upper().replace(" ", ""))
    if multiplier is None or not math.isfinite(amount) or amount <= 0:
        return None
    annual = amount * multiplier
    return annual if 1_000 <= annual <= 10_000_000 else None


def weighted_average(total, weight, digits=0):
    if not weight:
        return None
    return round(total / weight, digits)


def require_columns(column, required, path):
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"Missing columns in {path}: {', '.join(missing)}")


if __name__ == "__main__":
    main()
